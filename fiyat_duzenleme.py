#fiyat güncelleme sayfaı
#hasan kaan kartal
import tkinter as tk
from tkinter.ttk import Combobox
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk
from datetime import datetime
def on_button_click():
    secilen_isim = kutu.get()

    # Seçilen ismi Excel dosyasında arayarak kaçıncı satırda olduğunu bulma
    ws = wb["Sayfa1"]
    for satir in range(1, 3000):
        hucres_deger = (ws.cell(satir, 1).value)
        if hucres_deger == secilen_isim:
            ucret = float(ws.cell(satir, 6).value)
            e4.config(text=ucret)
            print("Seçilen değer", secilen_isim, "Excel'de", satir, ". satırda bulunuyor.")
            break
    else:
        e4.config(text="Seçilen değer Excel'de bulunamadı.")
        print("Seçilen değer Excel'de bulunamadı.")


def fiyat_ekle():
    eklenecek = float(ent1.get())

    try:
        eklenecek = float(eklenecek)  # Girilen değeri sayıya dönüştürmeye çalış
    except ValueError:
        e4.config(text="Lütfen geçerli bir sayı girin.")
        return

    # Hücre değerini güncelle
    secilen_isim = kutu.get()
    ws = wb["Sayfa1"]
    for satir in range(1, 3000):
        hucres_deger = ws.cell(satir, 1).value
        if hucres_deger == secilen_isim:
            eski_ucret = float(ws.cell(satir, 6).value)
            yeni_ucret = float(eski_ucret + eklenecek)  # Yeni ücreti eski ücrete ekle
            ws.cell(satir, 6, value=yeni_ucret)  # Excel hücresini güncelle
            ws.cell(satir, 8, value= datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            e4.config(text=yeni_ucret)
            print("Seçilen değer", secilen_isim, "Excel'de", satir, ". satırda bulunuyor.")
            break
    else:
        e4.config(text="Seçilen değer Excel'de bulunamadı.")
        print("Seçilen değer Excel'de bulunamadı.")

    # Yapılan değişiklikleri Excel dosyasına kaydet
    wb.save("C:/Users/hasan/Desktop/ilk_deneme.xlsx")


def fiyat_cikar():
    cikarilacak = float(ent1.get())

    try:
        cikarilacak = float(cikarilacak)  # Girilen değeri sayıya dönüştürmeye çalış
    except ValueError:
        e4.config(text="Lütfen geçerli bir sayı girin.")
        return

    # Hücre değerini güncelle
    secilen_isim = kutu.get()
    ws = wb["Sayfa1"]
    for satir in range(1, 3000):
        hucres_deger = ws.cell(satir, 1).value
        if hucres_deger == secilen_isim:
            eski_ucret = float(ws.cell(satir, 6).value)
            yeni_ucret = float(eski_ucret - cikarilacak)  # Yeni ücreti eski ücretten çıkar
            ws.cell(satir, 6, value=yeni_ucret)  # Excel hücresini güncelle
            ws.cell(satir, 8, value= datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            e4.config(text=yeni_ucret)
            print("Seçilen değer", secilen_isim, "Excel'de", satir, ". satırda bulunuyor.")
            break
    else:
        e4.config(text="Seçilen değer Excel'de bulunamadı.")
        print("Seçilen değer Excel'de bulunamadı.")

    # Yapılan değişiklikleri Excel dosyasına kaydet
    wb.save("C:/Users/hasan/Desktop/ilk_deneme.xlsx")


wb = load_workbook("C:/Users/hasan/Desktop/ilk_deneme.xlsx")


i = 1
isimler = []
ws = wb["Sayfa1"]
while i < 3000:
    if type(ws.cell(i, 1).value) == str:
        isimler.append(ws.cell(i, 1).value)
    else:
        print("satir sayisi : ", i)
        break
    i += 1
print(isimler)

pencere = tk.Tk()
pencere.geometry('400x600')
# ------------ Arka Plan-----------------
image = Image.open("KARTALlll.png")
photo = ImageTk.PhotoImage(image)
label = tk.Label(pencere, image=photo)
label.pack()
kutu = Combobox(pencere, values=isimler[1:])
kutu.place(x=150,y=20)

e1 = tk.Label(text="ÖDEME EKLE-ÇIKAR", font="Arial 12 bold")
e1.place(x=140, y=145)

buton = tk.Button(pencere, text="Seçimi Al", command=on_button_click)
buton.place(x=190,y=60)

buton_ekle = tk.Button(pencere, text="Ekle", command=fiyat_ekle)
buton_ekle.place(x=200,y=190)

buton_cikar = tk.Button(pencere, text="Çıkar", command=fiyat_cikar)
buton_cikar.place(x=197,y=230)

e4 = tk.Label(text="----", font="Arial 12 bold", fg="red")
e4.place(x=200, y=270)

ent1 = tk.Entry(width=30)
ent1.place(x=135, y=100)

pencere.mainloop()